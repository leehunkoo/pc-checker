#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, json, sys, argparse
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SECTIONS = [
    {"id":1,"title":"계정 및 비밀번호 관리","items":[
        {"id":"1-1","text":"개인 계정과 공용 계정을 분리하여 사용하고 있다."},
        {"id":"1-2","text":"계정 비밀번호는 8자리 이상(대소문자+숫자+특수문자 조합)으로 설정하여 있다."},
        {"id":"1-3","text":"동일 비밀번호를 타 시스템에서도 사용하지 않는다."},
        {"id":"1-4","text":"비밀번호를 주기적으로 변경하고 있다."},
        {"id":"1-5","text":"화면보호기 잠금은 10분 이내로 설정하여 있다."},
        {"id":"1-6","text":"자리비움·자리이동 시 화면 잠금을 실시한다."},
    ]},
    {"id":2,"title":"악성코드 및 보안프로그램 대응","items":[
        {"id":"2-1","text":"백신 프로그램이 설치되어 있다."},
        {"id":"2-2","text":"백신 엔진 및 패턴이 항상 최신버전으로 업데이트되어 있다."},
        {"id":"2-3","text":"실시간 감시 기능이 활성화되어 있다."},
        {"id":"2-4","text":"정기적(주 1회 이상) 정밀 검사를 실시하고 있다."},
        {"id":"2-5","text":"출처불명·의심스런 실행파일을 실행하지 않는다."},
        {"id":"2-6","text":"방화벽이 활성화되어 있다. (V3 또는 Windows 방화벽)"},
    ]},
    {"id":3,"title":"운영체제 및 소프트웨어 보안업데이트","items":[
        {"id":"3-1","text":"운영체제(OS) 자동 업데이트가 설정되어 있다."},
        {"id":"3-2","text":"주요 소프트웨어(브라우저, MS Office, 한글 등)가 최신버전이다."},
        {"id":"3-3","text":"보안 취약점을 발견하면 즉시 업데이트한다."},
    ]},
    {"id":4,"title":"개인정보 및 중요정보 보호","items":[
        {"id":"4-1","text":"개인정보 파일을 암호화하여 보관한다."},
        {"id":"4-2","text":"중요 문서는 책상 위 방치나 출력물을 방치하지 않는다."},
        {"id":"4-3","text":"개인정보가 포함된 파일을 개인 이메일로 전송하지 않는다."},
        {"id":"4-4","text":"업무 종료 후 바탕화면·다운로드 폴더에 개인정보를 방치하지 않는다."},
        {"id":"4-5","text":"출력물은 즉시 회수하며, 불필요한 문서는 파쇄한다."},
    ]},
    {"id":5,"title":"외부 저장매체 관리","items":[
        {"id":"5-1","text":"USB 등 이동식 저장매체 사용 시 승인 절차를 따른다."},
        {"id":"5-2","text":"사용 전·후 악성코드 검사를 실시한다."},
        {"id":"5-3","text":"미인가 저장매체를 연결하지 않는다."},
        {"id":"5-4","text":"분실 방지를 위한 관리대장을 운영한다."},
    ]},
    {"id":6,"title":"네트워크 사용 보안","items":[
        {"id":"6-1","text":"공용 Wi-Fi 사용 시 업무자료 전송을 제한한다."},
        {"id":"6-2","text":"VPN 등 안전한 원격 접속 솔루션을 사용한다."},
        {"id":"6-3","text":"불필요한 파일공유 기능이 비활성화되어 있다."},
    ]},
    {"id":7,"title":"이메일 및 인터넷 사용 보안","items":[
        {"id":"7-1","text":"출처가 불분명한 이메일 첨부파일을 열지 않는다."},
        {"id":"7-2","text":"피싱 의심 메일은 즉시 보안담당자에게 신고한다."},
        {"id":"7-3","text":"업무와 무관한 사이트 접속을 자제한다."},
    ]},
    {"id":8,"title":"물리적 보안","items":[
        {"id":"8-1","text":"PC에 자산관리 스티커가 부착되어 있다."},
        {"id":"8-2","text":"사무실 외부 반출 시 승인 절차를 따른다."},
        {"id":"8-3","text":"장비 폐기 시 저장매체 완전삭제를 실시한다."},
    ]},
    {"id":9,"title":"개인정보 보호","items":[
        {"id":"9-1","text":"개인정보 침해 사고 발생 시 즉시 보안담당자에게 신고한다."},
        {"id":"9-2","text":"악성코드 감염 의심 시 네트워크를 차단하고 신고한다."},
    ]},
]

def generate_pdf(results, inspector, filename):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_paths = [
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\gulim.ttc",
        r"C:\Windows\Fonts\NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    ]
    korean_font = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("KoreanFont", fp))
                korean_font = "KoreanFont"
                break
            except:
                continue

    doc = SimpleDocTemplate(filename, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=18*mm, bottomMargin=18*mm)

    def sty(name, **kw):
        return ParagraphStyle(name, fontName=korean_font, wordWrap="CJK", **kw)

    story = []
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("PC 보안 자가점검 보고서", sty("T", fontSize=20, textColor=colors.HexColor("#1e3a5f"), leading=28, spaceAfter=4)))
    story.append(Paragraph("정보보안 자율점검 및 보안취약점 개선을 위한 자가점검 결과", sty("S", fontSize=11, textColor=colors.HexColor("#64748b"), leading=16, spaceAfter=2)))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#3b82f6"), spaceAfter=8))

    info_data = [
        ["점검자", inspector.get("name","-"), "부서", inspector.get("dept","-")],
        ["점검일", inspector.get("date", datetime.now().strftime("%Y-%m-%d")), "OS", inspector.get("os","Windows")],
    ]
    info_table = Table(info_data, colWidths=[25*mm,55*mm,25*mm,55*mm])
    info_table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),korean_font),("FONTSIZE",(0,0),(-1,-1),9),
        ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#f1f5f9")),
        ("BACKGROUND",(2,0),(2,-1),colors.HexColor("#f1f5f9")),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),
        ("PADDING",(0,0),(-1,-1),6),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(info_table)
    story.append(Spacer(1,8*mm))

    total = sum(len(s["items"]) for s in SECTIONS)
    pass_c = sum(1 for r in results.values() if r.get("status")=="pass")
    warn_c = sum(1 for r in results.values() if r.get("status")=="warn")
    fail_c = sum(1 for r in results.values() if r.get("status")=="fail")
    skip_c = total - len(results)
    score  = round(pass_c/total*100) if total>0 else 0
    status_color = colors.HexColor("#22c55e") if score>=80 else colors.HexColor("#f59e0b") if score>=60 else colors.HexColor("#ef4444")

    story.append(Paragraph("점검 요약", sty("ST", fontSize=13, textColor=colors.HexColor("#1e3a5f"), leading=20, spaceBefore=4, spaceAfter=4)))
    sum_data = [
        ["전체 항목", f"{total}개", "점검 완료", f"{len(results)}개"],
        ["적합(양호)", f"{pass_c}개", "주의", f"{warn_c}개"],
        ["부적합", f"{fail_c}개", "미점검", f"{skip_c}개"],
        ["보안 점수", f"{score}점 / 100점", "등급", "양호" if score>=80 else "보통" if score>=60 else "부적합"],
    ]
    sum_table = Table(sum_data, colWidths=[35*mm,45*mm,35*mm,45*mm])
    sum_table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),korean_font),("FONTSIZE",(0,0),(-1,-1),9.5),
        ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#eff6ff")),
        ("BACKGROUND",(2,0),(2,-1),colors.HexColor("#eff6ff")),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#bfdbfe")),
        ("PADDING",(0,0),(-1,-1),7),
        ("FONTSIZE",(1,3),(1,3),11),("TEXTCOLOR",(1,3),(1,3),status_color),
    ]))
    story.append(sum_table)
    story.append(Spacer(1,8*mm))

    story.append(Paragraph("세부 점검 결과", sty("DT", fontSize=13, textColor=colors.HexColor("#1e3a5f"), leading=20, spaceBefore=4, spaceAfter=6)))

    STATUS_LABEL = {"pass":"적합","fail":"부적합","warn":"주의","manual":"수동확인","pending":"미점검"}
    STATUS_COLOR = {"pass":colors.HexColor("#dcfce7"),"fail":colors.HexColor("#fee2e2"),"warn":colors.HexColor("#fef9c3"),"manual":colors.HexColor("#f3e8ff"),"pending":colors.HexColor("#f8fafc")}
    STATUS_TEXT  = {"pass":colors.HexColor("#166534"),"fail":colors.HexColor("#991b1b"),"warn":colors.HexColor("#92400e"),"manual":colors.HexColor("#6b21a8"),"pending":colors.HexColor("#64748b")}

    for sec in SECTIONS:
        sec_data = [["번호","점검 항목","결과","세부 내용"]]
        for item in sec["items"]:
            r = results.get(item["id"],{})
            st = r.get("status","pending")
            detail = (r.get("detail") or "-")[:80]
            sec_data.append([
                Paragraph(item["id"], sty(f"i{item['id']}a", fontSize=8, leading=11)),
                Paragraph(item["text"], sty(f"i{item['id']}b", fontSize=8.5, leading=13)),
                Paragraph(STATUS_LABEL.get(st,"-"), sty(f"i{item['id']}c", fontSize=8, leading=11)),
                Paragraph(detail, sty(f"i{item['id']}d", fontSize=7.5, leading=11)),
            ])
        sec_table = Table(sec_data, colWidths=[12*mm,88*mm,18*mm,40*mm], repeatRows=1)
        ts_list = [
            ("FONTNAME",(0,0),(-1,-1),korean_font),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#e2e8f0")),
            ("PADDING",(0,0),(-1,-1),5),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ALIGN",(2,0),(2,-1),"CENTER"),
        ]
        for ri, item in enumerate(sec["items"],1):
            r = results.get(item["id"],{})
            st = r.get("status","pending")
            ts_list.append(("BACKGROUND",(2,ri),(2,ri),STATUS_COLOR.get(st,colors.white)))
            ts_list.append(("TEXTCOLOR",(2,ri),(2,ri),STATUS_TEXT.get(st,colors.HexColor("#334155"))))
        sec_table.setStyle(TableStyle(ts_list))
        sec_header = Paragraph(f"{sec['title']}", sty(f"sh{sec['id']}", fontSize=11, textColor=colors.HexColor("#1e3a5f"), leading=18, spaceBefore=6, spaceAfter=3))
        story.append(KeepTogether([sec_header, sec_table]))
        story.append(Spacer(1,4*mm))

    story.append(Spacer(1,6*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cbd5e1")))
    story.append(Spacer(1,3*mm))
    from reportlab.lib.enums import TA_LEFT
    story.append(Paragraph(f"본 보고서는 PC 보안 자가점검 시스템에 의해 자동 생성되었습니다. 생성시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ParagraphStyle("foot2", fontName=korean_font, fontSize=8, textColor=colors.HexColor("#94a3b8"), leading=12, alignment=TA_LEFT)))

    doc.build(story)
    print(f"[OK] PDF 생성 완료: {filename}")

def generate_excel(results, inspector, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "점검 요약"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cs(ws, row, col, value, bold=False, bg=None, fg="000000", size=11, align="left", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="맑은 고딕", size=size, bold=bold, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.border = border
        return c

    ws1.merge_cells("A1:H1")
    cs(ws1,1,1,"PC 보안 자가점검 보고서",bold=True,bg="1e3a5f",fg="FFFFFF",size=16,align="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    cs(ws1,2,1,"정보보안 자율점검 및 보안취약점 개선을 위한 자가점검 결과",bg="eff6ff",fg="3b82f6",size=10,align="center")

    info = ["점검자", inspector.get("name","-"), "부서", inspector.get("dept","-"), "점검일", inspector.get("date",datetime.now().strftime("%Y-%m-%d")), "OS", inspector.get("os","Windows")]
    for ci, v in enumerate(info,1):
        cs(ws1,4,ci,v,bold=ci%2==1,bg="f1f5f9" if ci%2==1 else None,fg="475569" if ci%2==1 else "1e293b",size=10,align="center")
    ws1.row_dimensions[4].height = 22

    total = sum(len(s["items"]) for s in SECTIONS)
    pass_c = sum(1 for r in results.values() if r.get("status")=="pass")
    warn_c = sum(1 for r in results.values() if r.get("status")=="warn")
    fail_c = sum(1 for r in results.values() if r.get("status")=="fail")
    skip_c = total - len(results)
    score  = round(pass_c/total*100) if total>0 else 0

    ws1.merge_cells("A6:H6")
    cs(ws1,6,1,"전체 점검 결과 요약",bold=True,bg="f8fafc",fg="1e3a5f",size=12)
    ws1.row_dimensions[6].height = 24

    sum_rows = [("전체 항목",f"{total}개","점검 완료",f"{len(results)}개"),("적합(양호)",f"{pass_c}개","주의",f"{warn_c}개"),("부적합",f"{fail_c}개","미점검",f"{skip_c}개"),("보안 점수",f"{score}점/100점","등급","양호" if score>=80 else "보통" if score>=60 else "부적합")]
    for ri, row in enumerate(sum_rows,7):
        for ci,(lbl,val) in enumerate([(row[0],row[1]),(row[2],row[3])],1):
            ws1.merge_cells(start_row=ri,start_column=(ci-1)*4+1,end_row=ri,end_column=(ci-1)*4+2)
            ws1.merge_cells(start_row=ri,start_column=(ci-1)*4+3,end_row=ri,end_column=(ci-1)*4+4)
            cs(ws1,ri,(ci-1)*4+1,lbl,bold=True,bg="eff6ff",fg="475569",size=10,align="center")
            cs(ws1,ri,(ci-1)*4+3,val,bold=True,fg="1e293b",size=11,align="center")
        ws1.row_dimensions[ri].height = 22

    ws2 = wb.create_sheet("세부 점검 결과")
    ws2.merge_cells("A1:G1")
    cs(ws2,1,1,"세부 점검 항목 결과",bold=True,bg="1e3a5f",fg="FFFFFF",size=14,align="center")
    ws2.row_dimensions[1].height = 30

    for ci,h in enumerate(["섹션","항목ID","점검 항목","점검 방법","결과","세부내용","비고"],1):
        cs(ws2,2,ci,h,bold=True,bg="334155",fg="FFFFFF",size=9,align="center")
    ws2.row_dimensions[2].height = 22

    SL = {"pass":"적합","fail":"부적합","warn":"주의","manual":"수동확인","pending":"미점검"}
    SB = {"pass":"dcfce7","fail":"fee2e2","warn":"fef9c3","manual":"f3e8ff","pending":"f8fafc"}
    SF = {"pass":"166534","fail":"991b1b","warn":"92400e","manual":"6b21a8","pending":"64748b"}

    row = 3
    for sec in SECTIONS:
        for item in sec["items"]:
            r = results.get(item["id"],{})
            st = r.get("status","pending")
            vals = [sec["title"],item["id"],item["text"],"자동" if r.get("autoChecked") else "수동",SL.get(st,"-"),r.get("detail","-"),r.get("notes","-")]
            for ci,v in enumerate(vals,1):
                c = ws2.cell(row=row,column=ci,value=v)
                c.font = Font(name="맑은 고딕",size=9)
                c.alignment = Alignment(horizontal="center" if ci in [1,2,4,5] else "left",vertical="center",wrap_text=True)
                c.border = border
                if ci==5:
                    c.fill = PatternFill("solid",fgColor=SB.get(st,"f8fafc"))
                    c.font = Font(name="맑은 고딕",size=9,bold=True,color=SF.get(st,"000000"))
                elif row%2==0:
                    c.fill = PatternFill("solid",fgColor="f9fafb")
            ws2.row_dimensions[row].height = 36
            row += 1

    for col,w in [(1,18),(2,8),(3,50),(4,10),(5,10),(6,35),(7,15)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(filename)
    print(f"[OK] Excel 생성 완료: {filename}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--format", choices=["pdf","excel","both"], default="pdf")
    parser.add_argument("--results", default=None)
    parser.add_argument("--name", default="홍길동")
    parser.add_argument("--dept", default="정보보호팀")
    parser.add_argument("--os", default="Windows")
    parser.add_argument("--output-dir", default=".")
    args = parser.parse_args()

    if args.results and os.path.exists(args.results):
        with open(args.results,"r",encoding="utf-8") as f:
            data = json.load(f)
        results = data.get("results", data)
        inspector_data = data.get("inspector", {})
    else:
        results = {}
        inspector_data = {}

    inspector = {
        "name": args.name or inspector_data.get("name","홍길동"),
        "dept": args.dept or inspector_data.get("dept","정보보호팀"),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "os": args.os or inspector_data.get("os","Windows"),
    }

    os.makedirs(args.output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.format in ("pdf","both"):
        generate_pdf(results, inspector, os.path.join(args.output_dir, f"보안점검보고서_{ts}.pdf"))
    if args.format in ("excel","both"):
        generate_excel(results, inspector, os.path.join(args.output_dir, f"보안점검보고서_{ts}.xlsx"))

if __name__ == "__main__":
    main()
