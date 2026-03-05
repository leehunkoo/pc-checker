#!/usr/bin/env python3
"""
PC 보안 자가점검 보고서 생성기
정보통신망 이용촉진 및 정보보호 등에 관한 법률 기반
사용법: python report_generator.py --format pdf|excel --results results.json
"""

import os
import json
import sys
import argparse
from datetime import datetime

# Windows CP949 환경에서 UTF-8 출력 강제
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─── REPORT DATA ─────────────────────────────────────────────────────────────
SECTIONS = [
    {"id": 1, "emoji": "🔐", "title": "계정 및 접근통제", "items": [
        {"id": "1-1", "text": "개인 계정과 공용 계정을 구분하여 사용하고 있다."},
        {"id": "1-2", "text": "계정 비밀번호는 8자리 이상(영문·숫자·특수문자 조합)으로 설정되어 있다."},
        {"id": "1-3", "text": "동일 비밀번호를 타 시스템과 중복 사용하지 않는다."},
        {"id": "1-4", "text": "비밀번호를 주기적으로 변경하고 있다."},
        {"id": "1-5", "text": "화면보호기 잠금이 10분 이내 자동 설정되어 있다."},
        {"id": "1-6", "text": "퇴근·이석 시 반드시 로그아웃 또는 화면 잠금을 실시한다."},
    ]},
    {"id": 2, "emoji": "🛡️", "title": "악성코드 및 보안프로그램 관리", "items": [
        {"id": "2-1", "text": "백신 프로그램이 설치되어 있다."},
        {"id": "2-2", "text": "백신 엔진 및 패턴이 최신 버전으로 업데이트되어 있다."},
        {"id": "2-3", "text": "실시간 감시 기능이 활성화되어 있다."},
        {"id": "2-4", "text": "정기적(주 1회 이상) 전체 검사를 실시하고 있다."},
        {"id": "2-5", "text": "불법·출처 불명 소프트웨어를 설치하지 않는다."},
        {"id": "2-6", "text": "방화벽이 활성화되어 있다. (V3 또는 Windows 방화벽)"},
    ]},
    {"id": 3, "emoji": "⚙️", "title": "운영체제 및 소프트웨어 보안패치", "items": [
        {"id": "3-1", "text": "운영체제(OS) 자동 업데이트가 설정되어 있다."},
        {"id": "3-2", "text": "주요 프로그램(한글, MS Office, 브라우저 등)이 최신 버전이다."},
        {"id": "3-3", "text": "보안 취약점 경고 발생 시 즉시 조치한다."},
    ]},
    {"id": 4, "emoji": "🔒", "title": "개인정보 및 중요정보 보호", "items": [
        {"id": "4-1", "text": "개인정보 파일은 암호화하여 저장한다."},
        {"id": "4-2", "text": "중요 문서는 사내 승인된 저장매체 또는 서버에 보관한다."},
        {"id": "4-3", "text": "개인정보가 포함된 파일을 개인 이메일로 전송하지 않는다."},
        {"id": "4-4", "text": "업무 종료 후 바탕화면·다운로드 폴더에 개인정보를 방치하지 않는다."},
        {"id": "4-5", "text": "출력물은 즉시 회수하며, 불필요한 문서는 파쇄한다."},
    ]},
    {"id": 5, "emoji": "💾", "title": "외부 저장매체 관리", "items": [
        {"id": "5-1", "text": "USB 등 이동식 저장매체 사용 시 승인 절차를 따른다."},
        {"id": "5-2", "text": "사용 전·후 악성코드 검사를 실시한다."},
        {"id": "5-3", "text": "미인가 저장매체를 연결하지 않는다."},
        {"id": "5-4", "text": "분실 방지를 위한 관리대장을 운영한다."},
    ]},
    {"id": 6, "emoji": "🌐", "title": "네트워크 사용 보안", "items": [
        {"id": "6-1", "text": "공용 Wi-Fi 사용 시 업무자료 열람을 자제한다."},
        {"id": "6-2", "text": "VPN 등 승인된 접속 방식만 사용한다."},
        {"id": "6-3", "text": "불필요한 파일공유 기능은 비활성화되어 있다."},
    ]},
    {"id": 7, "emoji": "📧", "title": "이메일 및 인터넷 사용 보안", "items": [
        {"id": "7-1", "text": "출처가 불분명한 이메일 첨부파일을 열지 않는다."},
        {"id": "7-2", "text": "피싱 의심 메일은 즉시 보안담당자에게 신고한다."},
        {"id": "7-3", "text": "업무와 무관한 사이트 접속을 자제한다. (DLP 통제 또는 수동 준수)"},
    ]},
    {"id": 8, "emoji": "🏢", "title": "물리적 보안", "items": [
        {"id": "8-1", "text": "PC에 자산관리 스티커가 부착되어 있다."},
        {"id": "8-2", "text": "사무실 외부 반출 시 승인 절차를 따른다."},
        {"id": "8-3", "text": "장비 폐기 시 저장매체 완전삭제(포맷이 아닌 완전삭제)를 실시한다."},
    ]},
    {"id": 9, "emoji": "🚨", "title": "사고 대응", "items": [
        {"id": "9-1", "text": "개인정보 유출 의심 시 즉시 보안담당자에게 보고한다."},
        {"id": "9-2", "text": "악성코드 감염 시 네트워크를 차단하고 신고한다."},
    ]},
]


# ─── PDF GENERATOR ─────────────────────────────────────────────────────────
def generate_pdf(results: dict, inspector: dict, filename: str = "pc_security_report.pdf"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, KeepTogether)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Windows 한글 폰트 우선 → Linux → macOS 순서
    font_paths = [
        # Windows 시스템 폰트
        r"C:\Windows\Fonts\malgun.ttf",       # 맑은 고딕 (Windows 기본)
        r"C:\Windows\Fonts\malgunbd.ttf",
        r"C:\Windows\Fonts\gulim.ttc",        # 굴림
        r"C:\Windows\Fonts\batang.ttc",       # 바탕
        r"C:\Windows\Fonts\NanumGothic.ttf",  # 나눔고딕 (설치된 경우)
        # Linux
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKkr-Regular.otf",
        # macOS
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    korean_font = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("KoreanFont", fp))
                korean_font = "KoreanFont"
                break
            except Exception:
                continue

    doc = SimpleDocTemplate(
        filename, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )

    W, H = A4
    styles = getSampleStyleSheet()

    def style(name, **kw):
        # wordWrap 기본 적용으로 칸 밖 삐져나옴 방지
        return ParagraphStyle(name, fontName=korean_font, wordWrap="CJK", **kw)

    title_style   = style("Title",   fontSize=20, textColor=colors.HexColor("#1e3a5f"), leading=28, spaceAfter=4)
    sub_style     = style("Sub",     fontSize=11, textColor=colors.HexColor("#64748b"), leading=16, spaceAfter=2)
    section_style = style("Section", fontSize=12, textColor=colors.HexColor("#1e293b"), leading=18)
    body_style    = style("Body",    fontSize=9.5, textColor=colors.HexColor("#334155"), leading=14)
    small_style   = style("Small",   fontSize=8, textColor=colors.HexColor("#94a3b8"), leading=12)

    story = []

    # ── Cover ──
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("PC 보안 자가점검 보고서", title_style))
    story.append(Paragraph("정보통신망 이용촉진 및 정보보호 등에 관한 법률 기반", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#3b82f6"), spaceAfter=8))

    # Inspector info table
    info_data = [
        ["점검자", inspector.get("name", "-"), "부서", inspector.get("dept", "-")],
        ["점검일", inspector.get("date", datetime.now().strftime("%Y-%m-%d")), "OS", inspector.get("os", "Windows/Linux")],
    ]
    info_table = Table(info_data, colWidths=[25*mm, 55*mm, 25*mm, 55*mm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), korean_font),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f1f5f9")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#475569")),
        ("TEXTCOLOR", (2,0), (2,-1), colors.HexColor("#475569")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ("PADDING", (0,0), (-1,-1), 6),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8*mm))

    # ── Summary ──
    total = sum(len(s["items"]) for s in SECTIONS)
    pass_c  = sum(1 for r in results.values() if r.get("status") == "pass")
    warn_c  = sum(1 for r in results.values() if r.get("status") == "warn")
    fail_c  = sum(1 for r in results.values() if r.get("status") == "fail")
    skip_c  = total - len(results)
    score   = round(pass_c / total * 100) if total > 0 else 0

    status_color = colors.HexColor("#22c55e") if score >= 80 else colors.HexColor("#f59e0b") if score >= 60 else colors.HexColor("#ef4444")
    
    story.append(Paragraph("점검 요약", style("SumTitle", fontSize=13, textColor=colors.HexColor("#1e3a5f"), leading=20, spaceBefore=4, spaceAfter=4)))
    
    sum_data = [
        ["총 항목", f"{total}개", "점검 완료", f"{len(results)}개"],
        ["양호(적합)", f"{pass_c}개", "주의", f"{warn_c}개"],
        ["미흡", f"{fail_c}개", "미점검", f"{skip_c}개"],
        ["보안 점수", f"{score}점 / 100점", "등급", "양호" if score >= 80 else "보통" if score >= 60 else "미흡"],
    ]
    sum_table = Table(sum_data, colWidths=[35*mm, 45*mm, 35*mm, 45*mm])
    sum_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), korean_font),
        ("FONTSIZE", (0,0), (-1,-1), 9.5),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#eff6ff")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#eff6ff")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#bfdbfe")),
        ("PADDING", (0,0), (-1,-1), 7),
        ("FONTSIZE", (1,3), (1,3), 11),
        ("TEXTCOLOR", (1,3), (1,3), status_color),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 8*mm))

    # ── Detailed results ──
    story.append(Paragraph("세부 점검 결과", style("DetTitle", fontSize=13, textColor=colors.HexColor("#1e3a5f"), leading=20, spaceBefore=4, spaceAfter=6)))

    STATUS_LABEL = {"pass": "양호", "fail": "미흡", "warn": "주의", "manual": "수동확인", "pending": "미점검"}
    STATUS_COLOR = {
        "pass": colors.HexColor("#dcfce7"), "fail": colors.HexColor("#fee2e2"),
        "warn": colors.HexColor("#fef9c3"), "manual": colors.HexColor("#f3e8ff"),
        "pending": colors.HexColor("#f8fafc"),
    }
    STATUS_TEXT = {
        "pass": colors.HexColor("#166534"), "fail": colors.HexColor("#991b1b"),
        "warn": colors.HexColor("#92400e"), "manual": colors.HexColor("#6b21a8"),
        "pending": colors.HexColor("#64748b"),
    }

    for sec in SECTIONS:
        sec_data = [["번호", "점검 내용", "결과", "세부 내용"]]
        for item in sec["items"]:
            r = results.get(item["id"], {})
            st = r.get("status", "pending")
            detail_text = r.get("detail", "-") or "-"
            # 긴 텍스트는 Paragraph로 감싸서 자동 줄바꿈
            sec_data.append([
                Paragraph(item["id"], style(f"cell_{item['id']}_id", fontSize=8, leading=11)),
                Paragraph(item["text"], style(f"cell_{item['id']}_text", fontSize=8.5, leading=13)),
                Paragraph(STATUS_LABEL.get(st, "-"), style(f"cell_{item['id']}_st", fontSize=8, leading=11)),
                Paragraph(detail_text[:80], style(f"cell_{item['id']}_detail", fontSize=7.5, leading=11)),
            ])

        # 칼럼 너비: 번호12 / 내용90 / 결과18 / 세부38 = 158mm (A4 본문 약 180mm)
        col_w = [12*mm, 88*mm, 18*mm, 40*mm]
        sec_table = Table(sec_data, colWidths=col_w, repeatRows=1)
        
        table_style = [
            ("FONTNAME", (0,0), (-1,-1), korean_font),
            ("FONTSIZE", (0,0), (-1,-1), 8.5),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
            ("PADDING", (0,0), (-1,-1), 5),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (2,0), (2,-1), "CENTER"),
        ]
        # Color each result row
        for ri, item in enumerate(sec["items"], 1):
            r = results.get(item["id"], {})
            st = r.get("status", "pending")
            bg = STATUS_COLOR.get(st, colors.white)
            tc = STATUS_TEXT.get(st, colors.HexColor("#334155"))
            table_style.append(("BACKGROUND", (2, ri), (2, ri), bg))
            table_style.append(("TEXTCOLOR", (2, ri), (2, ri), tc))
            table_style.append(("FONTSIZE", (2, ri), (2, ri), 8))

        sec_table.setStyle(TableStyle(table_style))

        sec_header = Paragraph(f"{sec['title']}", style(
            f"SH{sec['id']}", fontSize=11, textColor=colors.HexColor("#1e3a5f"),
            leading=18, spaceBefore=6, spaceAfter=3
        ))
        story.append(KeepTogether([sec_header, sec_table]))
        story.append(Spacer(1, 4*mm))

    # ── Improvements ──
    fail_items = [(s, i, results.get(i["id"], {})) for s in SECTIONS for i in s["items"]
                  if results.get(i["id"], {}).get("status") in ("fail", "warn")]
    
    if fail_items:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("개선 필요 항목", style("ImpTitle", fontSize=13, textColor=colors.HexColor("#991b1b"), leading=20, spaceBefore=4, spaceAfter=6)))

        imp_data = [["항목", "점검 내용", "상태", "조치 방법"]]
        improve_tips = {
            "2-4": "V3 전체검사 예약 설정 > 주 1회 이상 권장",
            "3-1": "설정 > Windows 업데이트 > 고급 옵션 > 자동 업데이트 활성화",
            "6-3": "제어판 > 네트워크 및 공유 센터 > 고급 공유 설정 > 파일 공유 비활성화",
        }
        for sec, item, r in fail_items:
            tip = improve_tips.get(item["id"], r.get("detail", "담당자 문의"))
            imp_data.append([
                Paragraph(item["id"], style(f"imp_id_{item['id']}", fontSize=8, leading=11)),
                Paragraph(item["text"], style(f"imp_text_{item['id']}", fontSize=8, leading=12)),
                Paragraph(STATUS_LABEL.get(r.get("status",""), "-"), style(f"imp_st_{item['id']}", fontSize=8, leading=11)),
                Paragraph(tip[:80], style(f"imp_tip_{item['id']}", fontSize=7.5, leading=11)),
            ])

        imp_table = Table(imp_data, colWidths=[12*mm, 68*mm, 18*mm, 60*mm])
        imp_table.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), korean_font),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#991b1b")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#fecaca")),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#fff7f7")),
            ("PADDING", (0,0), (-1,-1), 5),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(imp_table)

    # Footer note
    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cbd5e1")))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "본 보고서는 정보통신망 이용촉진 및 정보보호 등에 관한 법률에 근거한 개인 PC 보안 자가점검 결과입니다. "
        f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        small_style
    ))

    doc.build(story)
    print(f"[OK] PDF 보고서 생성 완료: {filename}")
    return filename


# ─── EXCEL GENERATOR ─────────────────────────────────────────────────────────
def generate_excel(results: dict, inspector: dict, filename: str = "pc_security_report.xlsx"):
    import openpyxl
    from openpyxl.styles import (Font, Fill, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart

    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "점검 요약"

    def cell_style(ws, row, col, value, bold=False, bg=None, fg="000000", size=11, align="left", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="맑은 고딕", size=size, bold=bold, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws1.merge_cells("A1:H1")
    c = cell_style(ws1, 1, 1, "개인 PC 보안 자가점검 보고서", bold=True, bg="1e3a5f", fg="FFFFFF", size=16, align="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    cell_style(ws1, 2, 1, "정보통신망 이용촉진 및 정보보호 등에 관한 법률 기반", bg="eff6ff", fg="3b82f6", size=10, align="center")

    # Inspector info
    ws1.merge_cells("A3:H3")
    ws1.row_dimensions[3].height = 8

    info = [
        ["점검자", inspector.get("name","-"), "부서", inspector.get("dept","-"), "점검일", inspector.get("date", datetime.now().strftime("%Y-%m-%d")), "OS", inspector.get("os","Windows/Linux")],
    ]
    for ci, v in enumerate(info[0], 1):
        is_label = ci % 2 == 1
        cell_style(ws1, 4, ci, v, bold=is_label, bg="f1f5f9" if is_label else None, fg="475569" if is_label else "1e293b", size=10, align="center")

    ws1.row_dimensions[4].height = 22

    # Summary stats
    total = sum(len(s["items"]) for s in SECTIONS)
    pass_c = sum(1 for r in results.values() if r.get("status") == "pass")
    warn_c = sum(1 for r in results.values() if r.get("status") == "warn")
    fail_c = sum(1 for r in results.values() if r.get("status") == "fail")
    skip_c = total - len(results)
    score  = round(pass_c / total * 100) if total > 0 else 0

    ws1.merge_cells("A6:H6")
    cell_style(ws1, 6, 1, "▣ 점검 결과 요약", bold=True, bg="f8fafc", fg="1e3a5f", size=12)
    ws1.row_dimensions[6].height = 24

    summary_rows = [
        ("총 점검 항목", f"{total}개", "점검 완료", f"{len(results)}개"),
        ("양호(적합)", f"{pass_c}개", "주의", f"{warn_c}개"),
        ("미흡", f"{fail_c}개", "미점검", f"{skip_c}개"),
        ("보안 점수", f"{score}점/100점", "등급", "양호" if score>=80 else "보통" if score>=60 else "미흡"),
    ]
    for ri, row in enumerate(summary_rows, 7):
        for ci, (lbl, val) in enumerate([(row[0],row[1]),(row[2],row[3])], 1):
            ws1.merge_cells(start_row=ri, start_column=(ci-1)*4+1, end_row=ri, end_column=(ci-1)*4+2)
            ws1.merge_cells(start_row=ri, start_column=(ci-1)*4+3, end_row=ri, end_column=(ci-1)*4+4)
            cell_style(ws1, ri, (ci-1)*4+1, lbl, bold=True, bg="eff6ff", fg="475569", size=10, align="center")
            val_cell = cell_style(ws1, ri, (ci-1)*4+3, val, bold=True, fg="1e293b", size=11, align="center")
            if ri == 10 and ci == 1:  # Score
                val_cell.font = Font(name="맑은 고딕", size=14, bold=True,
                                    color="166534" if score>=80 else "92400e" if score>=60 else "991b1b")
        ws1.row_dimensions[ri].height = 22

    # Section bar chart data
    ws1.merge_cells("A12:H12")
    cell_style(ws1, 12, 1, "▣ 분야별 점검 결과", bold=True, bg="f8fafc", fg="1e3a5f", size=12)
    ws1.row_dimensions[12].height = 24

    sec_headers = ["분야", "총 항목", "양호", "주의", "미흡", "미점검", "달성률(%)"]
    for ci, h in enumerate(sec_headers, 1):
        cell_style(ws1, 13, ci, h, bold=True, bg="1e3a5f", fg="FFFFFF", size=9, align="center")

    for ri, sec in enumerate(SECTIONS, 14):
        items = sec["items"]
        sp = sum(1 for i in items if results.get(i["id"],{}).get("status") == "pass")
        sw = sum(1 for i in items if results.get(i["id"],{}).get("status") == "warn")
        sf = sum(1 for i in items if results.get(i["id"],{}).get("status") == "fail")
        sk = len(items) - sp - sw - sf
        pct = round(sp/len(items)*100)
        row_vals = [sec["title"], len(items), sp, sw, sf, sk, pct]
        bg_color = "dcfce7" if pct >= 80 else "fef9c3" if pct >= 60 else "fee2e2"
        for ci, v in enumerate(row_vals, 1):
            cell_style(ws1, ri, ci, v, size=9.5, align="center",
                      bg=bg_color if ci == 7 else ("f8fafc" if ri%2==0 else None))
        ws1.row_dimensions[ri].height = 20

    # Column widths
    for col, w in [(1,18),(2,8),(3,8),(4,8),(5,8),(6,8),(7,10)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Detail ──
    ws2 = wb.create_sheet("세부 점검 결과")
    
    ws2.merge_cells("A1:G1")
    cell_style(ws2, 1, 1, "세부 점검 항목 결과", bold=True, bg="1e3a5f", fg="FFFFFF", size=14, align="center")
    ws2.row_dimensions[1].height = 30

    headers = ["분야", "항목ID", "점검 내용", "점검 방법", "결과", "상세내용", "비고"]
    for ci, h in enumerate(headers, 1):
        cell_style(ws2, 2, ci, h, bold=True, bg="334155", fg="FFFFFF", size=9, align="center")
    ws2.row_dimensions[2].height = 22

    STATUS_LABEL = {"pass":"양호","fail":"미흡","warn":"주의","manual":"수동확인","pending":"미점검"}
    STATUS_BG    = {"pass":"dcfce7","fail":"fee2e2","warn":"fef9c3","manual":"f3e8ff","pending":"f8fafc"}
    STATUS_FG    = {"pass":"166534","fail":"991b1b","warn":"92400e","manual":"6b21a8","pending":"64748b"}
    AUTO_KEYS    = {"1-5","2-1","2-2","2-3","3-1","4-1","5-3","6-2","6-3"}

    row = 3
    for sec in SECTIONS:
        sec_start = row
        for item in sec["items"]:
            r = results.get(item["id"], {})
            st = r.get("status","pending")
            method = "자동" if item["id"] in AUTO_KEYS else "수동"
            vals = [sec["title"], item["id"], item["text"], method,
                    STATUS_LABEL.get(st,"-"), r.get("detail","-"), r.get("note","-")]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=row, column=ci, value=v)
                c.font = Font(name="맑은 고딕", size=9)
                c.alignment = Alignment(horizontal="center" if ci in [1,2,4,5] else "left",
                                       vertical="center", wrap_text=True)
                c.border = border
                if ci == 5:
                    c.fill = PatternFill("solid", fgColor=STATUS_BG.get(st,"f8fafc"))
                    c.font = Font(name="맑은 고딕", size=9, bold=True, color=STATUS_FG.get(st,"000000"))
                elif row % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="f9fafb")
            ws2.row_dimensions[row].height = 36
            row += 1

    col_widths2 = [(1,18),(2,8),(3,50),(4,10),(5,10),(6,30),(7,15)]
    for col, w in col_widths2:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Improvements ──
    ws3 = wb.create_sheet("개선 사항")

    ws3.merge_cells("A1:F1")
    cell_style(ws3, 1, 1, "개선 필요 항목 및 조치 방안", bold=True, bg="991b1b", fg="FFFFFF", size=14, align="center")
    ws3.row_dimensions[1].height = 30

    imp_headers = ["항목ID", "분야", "점검 내용", "현재 상태", "조치 방안", "완료 여부"]
    for ci, h in enumerate(imp_headers, 1):
        cell_style(ws3, 2, ci, h, bold=True, bg="334155", fg="FFFFFF", size=9, align="center")
    ws3.row_dimensions[2].height = 22

    improve_tips = {
        "1-5": "설정 > 개인 설정 > 잠금 화면에서 화면 시간 제한을 10분 이하로 설정",
        "2-1": "Windows Defender 또는 서드파티 백신 프로그램 즉시 설치",
        "2-2": "백신 프로그램을 열고 '업데이트' 또는 '엔진 업데이트' 실행",
        "2-3": "Windows Defender > '바이러스 및 위협 방지 설정' > '실시간 보호' 활성화",
        "3-1": "설정 > Windows 업데이트 > 고급 옵션에서 자동 업데이트 활성화",
        "4-1": "설정 > 개인정보 및 보안 > 장치 암호화에서 BitLocker 활성화",
        "5-3": "제어판 > 자동 실행에서 이동식 미디어 자동 실행 비활성화",
        "6-2": "설정 > 네트워크 및 인터넷 > VPN에서 승인된 VPN 연결 설정",
        "6-3": "제어판 > 네트워크 및 공유 센터 > 고급 공유 설정에서 파일 공유 비활성화",
    }

    row = 3
    fail_found = False
    for sec in SECTIONS:
        for item in sec["items"]:
            r = results.get(item["id"], {})
            st = r.get("status","pending")
            if st in ("fail","warn"):
                fail_found = True
                tip = improve_tips.get(item["id"], r.get("detail","담당자 문의 필요"))
                vals = [item["id"], sec["title"], item["text"], STATUS_LABEL.get(st,"-"), tip, "□ 미완료"]
                bg = "fee2e2" if st == "fail" else "fef9c3"
                for ci, v in enumerate(vals, 1):
                    c = ws3.cell(row=row, column=ci, value=v)
                    c.font = Font(name="맑은 고딕", size=9)
                    c.alignment = Alignment(horizontal="center" if ci in [1,2,4,6] else "left",
                                           vertical="center", wrap_text=True)
                    c.border = border
                    c.fill = PatternFill("solid", fgColor=bg if ci in [1,4] else "fffbeb" if st=="warn" else "fff7f7")
                ws3.row_dimensions[row].height = 40
                row += 1

    if not fail_found:
        ws3.merge_cells(f"A3:F3")
        cell_style(ws3, 3, 1, "🎉 개선이 필요한 항목이 없습니다. 모든 항목 양호!", bg="dcfce7", fg="166534", size=12, align="center", bold=True)
        ws3.row_dimensions[3].height = 36

    col_widths3 = [(1,8),(2,18),(3,45),(4,10),(5,50),(6,10)]
    for col, w in col_widths3:
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(filename)
    print(f"[OK] Excel 보고서 생성 완료: {filename}")
    return filename


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="PC 보안 자가점검 보고서 생성기")
    parser.add_argument("--format", choices=["pdf","excel","both"], default="both", help="출력 형식")
    parser.add_argument("--results", default=None, help="점검 결과 JSON 파일 경로")
    parser.add_argument("--name", default="홍길동", help="점검자 성명")
    parser.add_argument("--dept", default="정보보호팀", help="부서")
    parser.add_argument("--os", default="Windows 11", help="운영체제")
    parser.add_argument("--output-dir", default=".", help="출력 디렉토리")
    args = parser.parse_args()

    # Load or generate sample results
    if args.results and os.path.exists(args.results):
        with open(args.results, "r", encoding="utf-8") as f:
            results = json.load(f)
    else:
        # Demo results
        results = {
            "1-5": {"status":"pass","detail":"화면보호기 잠금 5분 설정 확인됨","autoChecked":True},
            "2-1": {"status":"pass","detail":"Windows Defender 활성화 확인됨","autoChecked":True},
            "2-2": {"status":"pass","detail":"바이러스 정의: 최신","autoChecked":True},
            "2-3": {"status":"pass","detail":"실시간 보호 활성 상태","autoChecked":True},
            "3-1": {"status":"warn","detail":"업데이트 자동설치 비활성 — 수동 확인 필요","autoChecked":True},
            "4-1": {"status":"fail","detail":"BitLocker 드라이브 암호화 비활성화","autoChecked":True},
            "5-3": {"status":"warn","detail":"이동식 미디어 자동실행 활성화됨","autoChecked":True},
            "6-2": {"status":"manual","detail":"VPN 연결 상태 수동 확인 필요","autoChecked":True},
            "6-3": {"status":"warn","detail":"네트워크 파일공유 활성 감지됨","autoChecked":True},
            # Manual items
            "1-1": {"status":"pass","autoChecked":False},
            "1-2": {"status":"pass","autoChecked":False},
            "1-3": {"status":"pass","autoChecked":False},
            "1-4": {"status":"warn","detail":"비밀번호 변경 90일 초과","autoChecked":False},
            "2-4": {"status":"pass","autoChecked":False},
            "7-1": {"status":"pass","autoChecked":False},
            "7-2": {"status":"pass","autoChecked":False},
            "8-1": {"status":"pass","autoChecked":False},
            "9-1": {"status":"pass","autoChecked":False},
            "9-2": {"status":"pass","autoChecked":False},
        }

    inspector = {
        "name": args.name,
        "dept": args.dept,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "os": args.os,
    }

    os.makedirs(args.output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.format in ("pdf","both"):
        generate_pdf(results, inspector, os.path.join(args.output_dir, f"보안점검보고서_{ts}.pdf"))

    if args.format in ("excel","both"):
        generate_excel(results, inspector, os.path.join(args.output_dir, f"보안점검보고서_{ts}.xlsx"))


if __name__ == "__main__":
    main()
